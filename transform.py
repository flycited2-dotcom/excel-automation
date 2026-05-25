import json
import math
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

BASE_DIR = Path(__file__).parent

EXCLUDE_SUPPLIER_GROUPS = {'Кондиционеры инверторные', 'Кондиционеры он/офф'}
GREE_BRANDS = {'DAICHI', 'DANTEX'}
_TYPE_RE = re.compile(r'\s*\bon[-/]off\b|\s*\binv(erter|ertor)?\b|\s*\bинвертор\b', re.IGNORECASE)

_SPLITHUB_SUBSECTION_ORDER = {
    'Медная труба': 1,
    'Изоляция': 2,
    'Кронштейны': 3,
    'Дренаж': 5,
    'Крепёж': 7,
}

# Кастомный порядок расходников (нумерация по выводу: медь→изоляция→кронштейны→ленты→дренаж→кабель→крепёж→фреон)
_RASHODNIKI_CUSTOM = [3,1,5,2,4,6,7,10,8,12,9,11,13,14,15,16,17,18,19,25,20,21,23,24,22]
_RASHODNIKI_SORT_KEY = {pos: rank for rank, pos in enumerate(_RASHODNIKI_CUSTOM, start=1)}


def _splithub_subsection_order(subsection, model_lower):
    order = _SPLITHUB_SUBSECTION_ORDER.get(subsection)
    if order is not None:
        return order
    if subsection == 'Прочее':
        if 'пвс' in model_lower or 'провод' in model_lower:
            return 6
        if 'фреон' in model_lower:
            return 8
        return 4
    return 0


def _splithub_group_from_series(series_lower):
    if any(k in series_lower for k in ['кассетная', 'напольно-потолочная']):
        return 'Полупромышленные сплит-системы'
    if any(k in series_lower for k in ['inverter', 'invertor', 'inv']):
        return 'Инверторные сплит-системы'
    if any(k in series_lower for k in ['on/off', 'on-off']):
        return 'Классические сплит-системы'
    return None


def _splithub_group_from_desc(desc_lower):
    if 'инвертор' in desc_lower:
        return 'Инверторные сплит-системы'
    if 'on/off' in desc_lower or 'on-off' in desc_lower:
        return 'Классические сплит-системы'
    return None




def read_splithub_price(filepath):
    df_raw = pd.read_excel(filepath, header=None)

    records = []
    current_brand = ''
    current_group = None
    current_section = ''
    current_subsection = ''

    for _, row in df_raw.iterrows():
        v0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        v1 = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        v2 = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        v3 = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
        v4 = row.iloc[4] if len(row) > 4 and pd.notna(row.iloc[4]) else None

        if v1 == 'ID' or v0 == 'Прайс-лист СплитХаб' or v0.startswith('Оптовые'):
            continue

        # Заголовок раздела/бренда: col0 есть, col1 пусто
        if v0 and not v1:
            if v0 == 'Медная труба':
                current_section = 'Медная труба'
                current_group = 'Расходные материалы для монтажа (медь)'
                current_subsection = 'Медная труба'
                current_brand = ''
            elif v0 == 'Расходники':
                current_section = 'Расходники'
                current_group = 'Расходные материалы для монтажа (медь)'
                current_subsection = ''
                current_brand = ''
            else:
                current_brand = v0
                current_section = 'split'
                current_group = None
                current_subsection = ''
            continue

        # Подзаголовок серии: col0 пусто, col1 есть
        if not v0 and v1 and v1 != 'Фото':
            if current_section == 'Расходники':
                current_subsection = v1
            elif current_section == 'split':
                current_group = _splithub_group_from_series(v1.lower())
            continue

        # Строка товара: col0 — число
        try:
            float(v0)
        except (ValueError, TypeError):
            continue

        if v4 is None:
            continue
        try:
            price = float(v4)
        except (ValueError, TypeError):
            continue

        group = current_group
        if current_section == 'split' and group is None:
            group = _splithub_group_from_desc(v3.lower())
        if not group:
            continue

        sub_order = 0
        if group == 'Расходные материалы для монтажа (медь)':
            sub_order = _splithub_subsection_order(current_subsection, v2.lower())

        clean_model = _TYPE_RE.sub('', v2).strip()
        name = f"{clean_model} {v3}".strip() if v3 else clean_model
        if current_brand in GREE_BRANDS:
            name = f"{name} (GREE)"
        records.append({
            'article': str(int(float(v1))) if v1 else v2,
            'group': group,
            'brand': current_brand,
            'name': name,
            'price_in': price,
            'price_out': price,
            'subsection_order': sub_order,
        })

    return pd.DataFrame(records)


COLORS = {
    'header_bg':   '1F4E79',
    'header_font': 'FFFFFF',
    'city_font':   'A0B4C8',
    'phone_font':  'D4E4F7',
    'tg_font':     '64B5F6',
    'group_bg':    'BDD7EE',
    'group_font':  '1F4E79',
    'alt_row':     'EBF3FB',
}

COL_WIDTHS = {'A': 7.57, 'B': 13.14, 'C': 22.5, 'D': 77.29, 'E': 16.0, 'F': 15.57}


def load_config():
    with open(BASE_DIR / 'config.json', encoding='utf-8') as f:
        return json.load(f)


def _side():
    return Side(style='thin', color='B8CCE4')


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _apply(cell, style):
    for attr, val in style.items():
        setattr(cell, attr, val)


def _table_px_width():
    # Точная ширина столбца в пикселях по формуле Excel (MDW=7 для шрифта по умолч.)
    # px = trunc(((256*width + trunc(128/MDW)) / 256) * MDW)
    mdw = 7
    pad = int(128 // mdw)
    return sum(int((256 * w + pad) / 256 * mdw) for w in COL_WIDTHS.values())


def _add_banner(ws, config, dark_fill):
    """Вставляет баннер-картинку первой строкой шапки во всю ширину таблицы.
    Возвращает смещение строк: 1 если баннер вставлен, иначе 0.

    banner_aspect (config): целевое соотношение ширина:высота. Если картинка
    «выше» (соотношение меньше), её центральная полоса обрезается до этого
    соотношения — баннер остаётся во всю ширину, но ниже. None = без обрезки."""
    import io
    from PIL import Image as PILImage

    banner_file = config.get('banner_file', 'banner.png')
    banner_path = BASE_DIR / banner_file
    if not banner_path.exists():
        return 0

    try:
        pil = PILImage.open(banner_path)
        w, h = pil.size
        target_aspect = config.get('banner_aspect')
        if target_aspect and (w / h) < target_aspect:
            new_h = int(round(w / target_aspect))
            top = (h - new_h) // 2
            pil = pil.crop((0, top, w, top + new_h))
            w, h = pil.size
        buf = io.BytesIO()
        pil.convert('RGB').save(buf, format='PNG')
        buf.seek(0)
        img = XLImage(buf)
    except Exception as e:
        print(f"  Предупреждение: не удалось загрузить баннер ({e})")
        return 0

    # Высота строки — по пропорциям картинки при ширине таблицы
    table_w = _table_px_width()
    banner_h_px = int(round(table_w * h / float(w)))

    for col in range(1, 7):
        ws.cell(1, col).fill = dark_fill
    ws.row_dimensions[1].height = banner_h_px * 0.75  # пиксели → пункты

    # Якорь точно по диапазону A1:F1 — правый край совпадает со столбцом F
    _from = AnchorMarker(col=0, colOff=0, row=0, rowOff=0)
    _to = AnchorMarker(col=6, colOff=0, row=1, rowOff=0)
    img.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=_to)
    ws.add_image(img)
    return 1


# ─── Чтение порядка групп из файла ────────────────────────────────────────────

def load_group_order(config):
    order_file = BASE_DIR / config.get('group_order_file', 'Места_по_группам_товара_исправлено.xlsx')
    if not order_file.exists():
        return {}

    df = pd.read_excel(order_file, skiprows=1, header=None)
    order = {}
    for _, row in df.iterrows():
        group_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        position = row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else None
        if group_name and group_name != 'Группа товара' and position is not None:
            try:
                order[group_name] = int(float(position))
            except (ValueError, TypeError):
                pass
    return order


# ─── Чтение наценок из файла ──────────────────────────────────────────────────

def load_markups(config):
    markups_path = BASE_DIR / config.get('markups_file', 'наценки по группам товара.xlsx')
    if not markups_path.exists():
        return {}, {}

    df = pd.read_excel(markups_path, skiprows=2, header=None)

    group_markups = {}
    homeline_prices = {}
    in_homeline = False

    for _, row in df.iterrows():
        col_b = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        col_c = row.iloc[2] if len(row) > 2 else None

        if col_b == 'Артикул':
            in_homeline = True
            continue

        if not in_homeline:
            if col_b and col_b not in ('Группа товара', 'Наценка (%)', 'nan'):
                try:
                    markup_val = int(float(col_c)) if pd.notna(col_c) else 0
                    group_markups[col_b] = markup_val
                except (ValueError, TypeError):
                    pass
        else:
            article = col_b
            col_d = row.iloc[3] if len(row) > 3 else None
            if article and article != 'nan' and col_d is not None and pd.notna(col_d):
                try:
                    homeline_prices[article] = float(col_d)
                except (ValueError, TypeError):
                    pass

    return group_markups, homeline_prices


# ─── Чтение прайса поставщика ─────────────────────────────────────────────────

def read_supplier_price(filepath, config):
    skiprows = config.get('read_options', {}).get('skiprows', 0)
    df = pd.read_excel(filepath, skiprows=skiprows, header=None)

    col_names = ['article', 'group', 'brand', 'name', 'price_in', 'stock', 'stock2', 'order']
    df.columns = col_names[:len(df.columns)]

    df = df[~df['price_in'].isin(['Цена', 'руб.'])]
    df = df.dropna(subset=['name', 'price_in'])
    df['price_in'] = pd.to_numeric(df['price_in'], errors='coerce')
    df = df.dropna(subset=['price_in'])
    df = df.reset_index(drop=True)

    df['article'] = df['article'].astype(str).str.strip()
    df['group'] = df['group'].astype(str).str.strip()
    df['brand'] = df['brand'].fillna('').astype(str).str.strip()
    df['name'] = df['name'].astype(str).str.strip()

    df = df[~df['group'].isin(EXCLUDE_SUPPLIER_GROUPS)].reset_index(drop=True)
    return df


# ─── Применение наценок ───────────────────────────────────────────────────────

def apply_pricing(df, group_markups, homeline_prices, default_markup):
    df = df.copy()
    prices_out = []

    has_price_out = 'price_out' in df.columns

    for _, row in df.iterrows():
        if has_price_out and pd.notna(row.get('price_out')):
            prices_out.append(row['price_out'])
            continue

        article = row['article']

        if article in homeline_prices:
            prices_out.append(homeline_prices[article])
            continue

        group = row['group']
        markup_pct = group_markups.get(group)
        if markup_pct is None:
            markup_pct = group_markups.get(group.rstrip())
        if markup_pct is None:
            markup_pct = default_markup

        price = row['price_in'] * (1 + markup_pct / 100)
        prices_out.append(round(price, 0))

    df['price_out'] = prices_out
    return df


# ─── Лист «Прайс клиента» ────────────────────────────────────────────────────

def build_pricelist(df, wb, config, group_order, group_after=None):
    ws = wb.active
    ws.title = 'Прайс клиента'

    for letter, w in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    dark_fill = PatternFill('solid', start_color=COLORS['header_bg'])

    # ─── Баннер (строка 1, если файл есть) ───
    # Контакты (название, город, телефоны) вшиты в сам баннер,
    # поэтому отдельных текстовых строк в шапке нет.
    off = _add_banner(ws, config, dark_fill)

    # ─── Заголовки таблицы ───
    header_row = 1 + off
    headers = ['№', 'Артикул', 'Бренд', 'Наименование', 'Цена (руб.)', 'Заказ (шт.)']
    header_style = {
        'font': Font(bold=True, size=10, name='Arial', color=COLORS['header_font']),
        'fill': PatternFill('solid', start_color=COLORS['header_bg']),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _border(),
    }
    for ci, h in enumerate(headers, 1):
        _apply(ws.cell(row=header_row, column=ci, value=h), header_style)

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:F{header_row}'

    # ─── Сортировка групп по заданному порядку ───
    max_order = 9999
    unique_groups = df['group'].unique().tolist()
    _group_after = group_after or {}
    _group_positions = config.get('group_positions', {})

    def group_sort_key(g):
        if g in _group_positions:
            return _group_positions[g]
        pos = group_order.get(g)
        if pos is None:
            pos = group_order.get(g.rstrip())
        if pos is None and g in _group_after:
            parent = _group_after[g]
            parent_pos = group_order.get(parent) or group_order.get(parent.rstrip())
            if parent_pos is not None:
                pos = parent_pos + 0.5
        return pos if pos is not None else max_order

    sorted_groups = sorted(unique_groups, key=group_sort_key)

    # ─── Данные ───
    row = header_row + 1
    num = 1

    for group in sorted_groups:
        gdf = df[df['group'] == group]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        gc = ws.cell(row=row, column=1, value=group)
        gc.font = Font(bold=True, size=10, name='Arial', color=COLORS['group_font'])
        gc.fill = PatternFill('solid', start_color=COLORS['group_bg'])
        gc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        row += 1

        for _, item in gdf.iterrows():
            is_alt = (num % 2 == 0)
            alt_fill = PatternFill('solid', start_color=COLORS['alt_row']) if is_alt else None

            vals = [num, item['article'], item['brand'], item['name'], item['price_out'], '']

            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = Font(size=10, name='Arial')
                cell.border = _border()
                cell.alignment = Alignment(vertical='center', wrap_text=(ci in (3, 4)))
                if ci == 5:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = '#,##0'
                if alt_fill:
                    cell.fill = alt_fill

            # Автовысота: максимум по колонкам C (22.5 шир) и D (77.29 шир)
            brand_lines = max(1, math.ceil(len(str(item['brand'])) / 24))
            name_lines  = max(1, math.ceil(len(str(item['name']))  / 85))
            ws.row_dimensions[row].height = max(brand_lines, name_lines) * 14.5

            row += 1
            num += 1


# ─── Главная функция ──────────────────────────────────────────────────────────

def transform(input_path, output_dir=None):
    config = load_config()

    print(f"  Читаю файл: {Path(input_path).name}")
    df = read_supplier_price(input_path, config)
    print(f"  Загружено позиций: {len(df)}")

    print("  Загружаю наценки...")
    group_markups, homeline_prices = load_markups(config)
    default_markup = config.get('default_markup', 10)
    print(f"  Наценки по группам: {len(group_markups)}")
    print(f"  HOMELINE фикс. цены: {len(homeline_prices)}")

    print("  Загружаю порядок групп...")
    group_order = load_group_order(config)
    print(f"  Позиций в порядке: {len(group_order)}")

    splithub_path = BASE_DIR / config.get('splithub_file', 'splithub-price.xlsx')
    if splithub_path.exists():
        print("  Загружаю прайс СплитХаб...")
        df_splithub = read_splithub_price(splithub_path)
        print(f"  СплитХаб позиций: {len(df_splithub)}")
        df = pd.concat([df, df_splithub], ignore_index=True)
    else:
        print(f"  Предупреждение: файл СплитХаб не найден ({splithub_path.name})")

    df = apply_pricing(df, group_markups, homeline_prices, default_markup)

    # Сортировка расходников: сначала по секциям, затем кастомный порядок
    if 'subsection_order' in df.columns:
        df = df.sort_values('subsection_order', kind='stable').reset_index(drop=True)
        mask_rash = df['group'] == 'Расходные материалы для монтажа (медь)'
        rash_idx = df[mask_rash].index.tolist()
        for seq_num, idx in enumerate(rash_idx, start=1):
            df.at[idx, 'subsection_order'] = _RASHODNIKI_SORT_KEY.get(seq_num, seq_num)
        df = df.sort_values('subsection_order', kind='stable').reset_index(drop=True)

    group_after = config.get('group_after', {})
    wb = openpyxl.Workbook()
    build_pricelist(df, wb, config, group_order, group_after)
    # Лист «Наценки» НЕ добавляется — только 1 лист «Прайс клиента»

    if output_dir is None:
        output_dir = str(BASE_DIR / 'output')
    os.makedirs(output_dir, exist_ok=True)

    prefix = config.get('output_prefix', 'БытТехОпт')
    ts = datetime.today().strftime('%Y%m%d')
    output_path = str(Path(output_dir) / f"{prefix}_{ts}.xlsx")
    wb.save(output_path)
    print(f"  Сохранено: {output_path}")
    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Использование: python transform.py <прайс_поставщика.xlsx>")
        sys.exit(1)
    transform(sys.argv[1])
